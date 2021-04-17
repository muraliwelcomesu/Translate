from selenium import webdriver
from selenium.webdriver.firefox.options import Options
import openpyxl,os,sys
import time
import io,re
import time
import subprocess
import Utils

class To_Dict():
    
    def excel_to_dict(self,ExcelName,lang):
        l_dict1 = {}
        l_dict_lang = {'en':'C','tn':'D','ml':'E','hi':'F'}
        wb1 =  openpyxl.load_workbook(ExcelName)
        sheet = wb1['Translate']
        for row in range(2,sheet.max_row + 1):
            if sheet['C' + str(row)].value is not None:
                l_index = int(sheet['A' + str(row)].value)
                l_dict1[l_index] = sheet[l_dict_lang[lang] + str(row)].value
        return l_dict1
        
    def excel_to_dict2(self,ExcelName):
        l_dict1 = {}
        wb1 =  openpyxl.load_workbook(ExcelName)
        sheet = wb1['Output_new']
        for row in range(2,sheet.max_row + 1):
            if sheet['B' + str(row)].value is not None:
                l_index = int(sheet['A' + str(row)].value)
                l_dict1[l_index] = sheet[l_dict_lang[lang] + str(row)].value
        return l_dict1 

class To_Excel():

    def format_list(self,p_list):
        l_space = '%20'
        l_newline = '%0A'
        l_str = ''
        l_str_original = ''
        for line in p_list:
            l_str_original = l_str_original + line + '\n'
            l_str = l_str + line.replace(' ',l_space) + l_newline
        return l_str,l_str_original


    def dict_to_Excel(self,ExcelName,dict_text):
        #l_dict_lang = {'en':'B','tn':'C','ml':'D','hi':'E'}
        wb1 =  openpyxl.load_workbook(ExcelName)
        sheet = wb1.create_sheet()
        sheet.title = 'Output_new'
        i = 1
        l_str = ''
        for key,values in dict_text.items():
            l_str = ''
            i = i + 1
            if int(len(values)) > 0:
                for k in values:
                    l_str = l_str + k
                sheet['A' + str(i)].value = key
                sheet['B' + str(i)].value = l_str
                
        wb1.save(ExcelName)
        print('Completed dict_to_Excel')
        
    def update_excel(self,ExcelName,p_dict):
        wb1 =  openpyxl.load_workbook(ExcelName)
        sheet = wb1['Translate']
        l_key_list = p_dict.keys()
        for keys in l_key_list:
            for row in range(2,sheet.max_row + 1):
                if int(keys) == int(sheet['A' + str(row)].value):
                    sheet['C' + str(row)].value = p_dict[keys]
                    break
        wb1.save(ExcelName)
        print('completed update_excel')
        
    def srt_to_Excel(self,FileName):
        print('Start of prepareExcel for File-',FileName)
        
        file = open(FileName,encoding="utf-8")
        filename1, file_extension1 = os.path.splitext(FileName)
        l_line_no = 0
        l_prev_line_no = -1
        l_dict1 = {}
        l_dict2 = {}
        l_tmp_list = []
        for line in file:
            try:
                    l_line_no = int(line)
                    if l_prev_line_no != l_line_no:
                        if len(l_tmp_list) > 0:
                            l_str,l_str_original = obj.format_list(l_tmp_list)
                            l_dict1[l_prev_line_no] = l_str
                            l_dict2[l_prev_line_no] = l_str_original
                            l_prev_line_no = l_line_no
                            l_tmp_list[:] = []
            except:
                if re.search('[a-zA-Z]', line) is not None:
                    if int(l_prev_line_no) <0:
                        l_prev_line_no = l_line_no
                    l_tmp_list.append(line)

        if len(l_tmp_list) > 0:
            l_str,l_str_original  = self.format_list(l_tmp_list)
            l_dict1[l_prev_line_no] = l_str
            l_dict2[l_prev_line_no] = l_str_original
            l_prev_line_no = l_line_no
            l_tmp_list[:] = []
            
        print('Dict preparation complete')
        if len(l_dict1) > 0:
            ExcelFileName = filename1 + '.xlsx'
            l_file_found = Utils.CheckFile(ExcelFileName)
            if not l_file_found:
                wb1 = openpyxl.Workbook()
                sheet = wb1.create_sheet()
                sheet.title = 'Translate'
                sheet['A1'].value  = 'Line No'
                sheet['B1'].value  = 'English Original'
                sheet['C1'].value  = 'English'
                wb1.save(ExcelFileName)
            wb1 = openpyxl.load_workbook(ExcelFileName)
            sheet = wb1['Translate']
            for key,value in l_dict1.items():
                l_row_num = sheet.max_row + 1;
                sheet['A' + str(l_row_num)].value = key
                sheet['B' + str(l_row_num)].value = l_dict2[key]
                sheet['C' + str(l_row_num)].value = value
            wb1.save(ExcelFileName)
        print('Completed Excel preparation')    

class Translation():
    def step1(self):
        print('Step1:Prepare ExcelFile for Each srt Files')
        work_dir = 'D:\\work_dir'
        FileNames = Utils.GetFilesList(work_dir,'srt')
        obj = To_Excel()
        for filename in FileNames:
            obj.srt_to_Excel(filename)
    
    def step2(self):
        print('Step2 - Translation')
        work_dir = 'D:\\work_dir'
        ExcelNames = Utils.GetFilesList(work_dir,'xlsx')
        for ExcelName in ExcelNames:
            obj = To_Dict()
            os.chdir('D:\\work_dir')
            l_dict1 = obj.excel_to_dict(ExcelName,'en')
            l_dict2 = l_dict1 
            print('Dict preparation complete for English text')
            l_tmp_list = []
            l_outlist = []
            l_count = 0
            #print(l_dict1)
            for key,value in l_dict1.items():
                l_tmp_list[:] = []
                l_tmp_list.append(value)
                l_tmp_cnt = 0
                while int(l_tmp_cnt) < 1:
                    print('inside loop calling translate')
                    l_count = l_count + 1
                    try:
                        l_output = obj.google_Translate('en','ta',l_tmp_list)
                    except:
                        print('Some Exception in translating')
                    if l_output is not None:
                        l_tmp_cnt = 2;
                    if l_count > 5:
                        print('retry again-'+str(l_count))
                        l_count = 0
                        break
                    l_outlist.append(l_output)
                    l_dict2[key] = l_output
                    time.sleep(1)
            print('done')
            return l_dict2    

            
        
            

if __name__ == "__main__":
    print('Starting')
    obj_translate = Translation()
    obj_translate.step1() #Prepare excel  
    l_dict = obj_translate.step2() #Translate  
        
    
    
    
    
